"""
ws_alpha.py — Stock Note 기반 종목 선정·매매 알고리즘
================================================================
15개월치 Stock Note에서 추출한 투자 원칙을 규칙 엔진으로 구현.

핵심 명제 (노트 원문 기반):
  "회사 실적, 점유율, 턴어라운드 체력, 브랜드가 괜찮은데
   주가는 안 올라가는 종목을 고르고 모아가면서 기다리면 나중에 오름"
   → 이것이 알파의 원천. 나머지 규칙은 이 명제를 지키기 위한 장치.

실행:
    python ws_alpha.py --demo              # 합성 데이터로 로직 검증
    python ws_alpha.py --run --date 20260810   # 실데이터 (pykrx + DART 필요)

의존:
    pip install pykrx opendartreader pandas numpy
    (Windows) set DART_API_KEY=...          (또는 setx DART_API_KEY "..." 로 영구 저장)
"""

from __future__ import annotations

import argparse
import os
from dataclasses import dataclass, field
from typing import Optional

import numpy as np
import pandas as pd


# ═══════════════════════════════════════════════════════════════
# 1. 설정 — 노트에서 추출한 임계값
# ═══════════════════════════════════════════════════════════════

@dataclass
class Config:
    # ---- 유니버스 (노트: "시총이 작아서 변동성 크다" → 하한 존재, 단 중소형주 선호)
    min_mktcap_eok: float = 800          # 시총 800억 이상
    max_mktcap_eok: float = 400_000      # 초대형주 제외 (삼전은 별도 코어 버킷)
    min_turnover_eok: float = 3.0        # 20일 평균 거래대금 3억 이상

    # ---- 하드 필터 (노트: "망하지 않을 회사"가 전제)
    max_debt_ratio: float = 200.0        # 부채비율 200% 초과 배제
                                         # (풀무원 300% = "아쉬움"으로 명시 → 감점 대상)
    min_roe: float = 5.0                 # ROE 5% 미만 배제 (KCC글라스 5% = 하한선)
    require_positive_op: bool = True     # 최근 4분기 누적 영업이익 > 0
    max_3m_return: float = 0.60          # 3개월 +60% 이상 = 테마 급등 → 신규진입 금지
                                         # ("인탑스 테마로 급등 → Exit", "삼성SDI 테마 → 팔 타임")

    # ---- 4대 팩터 가중치 (노트 언급 빈도 + 의사결정 영향력 기준)
    w_quality: float = 0.30              # 체력: ROE, 마진, 부채, 이익 안정성
    w_value: float = 0.28                # 가격: PER, PBR, 배당수익률
    w_gap: float = 0.27                  # ★괴리: 실적은 멀쩡한데 주가만 빠진 정도
    w_payout: float = 0.15               # 주주환원 여력: 낮은 배당성향 + 순현금

    # ---- 포지션 사이징 (노트: "한 종목 10% 이상 안 가져가기")
    max_weight_single: float = 0.10
    max_weight_sector: float = 0.25
    target_positions: int = 20           # "종목 수 5개 이하로 축소" 발언 있으나
                                         # 실제 보유는 40~60개 → 코어 20 + 위성으로 운용

    # ---- 매매 규칙
    entry_tranches: tuple = (0.4, 0.3, 0.3)   # 보초병 → 분할
    add_down_step: float = -0.20              # -20%마다 물타기 1회
    max_add_downs: int = 3                    # 최대 3회 (그 이상은 스토리 재검증)
    take_profit_recover_cost: float = 1.00    # +100% → 원금 회수, 잔량 장기보유
    trim_on_theme_spike: float = 0.50          # 3개월 +50% 급등 시 1/3 트림

    # ---- 섹터 틸트 (노트: "한국 특산품에 집중")
    sector_tilt: dict = field(default_factory=lambda: {
        "음식료": 0.06, "화장품": 0.06, "건강기능식품": 0.06, "제약바이오": 0.04,
        "반도체소재": 0.05, "조선기자재": 0.04, "방산": 0.04, "지주": 0.03,
        "물류": 0.02, "건설": -0.02, "게임": -0.02, "2차전지": -0.03,
        "증권": -0.02, "테마": -0.10,
    })


CFG = Config()


# ═══════════════════════════════════════════════════════════════
# 2. 유틸 — 횡단면 백분위 스코어
# ═══════════════════════════════════════════════════════════════

def pct_rank(s: pd.Series, ascending: bool = True) -> pd.Series:
    """결측은 중앙값(0.5) 처리. 0~1 백분위."""
    r = s.rank(pct=True, ascending=ascending, na_option="keep")
    return r.fillna(0.5)


def winsor(s: pd.Series, lo=0.01, hi=0.99) -> pd.Series:
    return s.clip(s.quantile(lo), s.quantile(hi))


# ═══════════════════════════════════════════════════════════════
# 3. 4대 팩터 스코어러
# ═══════════════════════════════════════════════════════════════

def score_quality(df: pd.DataFrame) -> pd.Series:
    """
    체력. 노트에서 가장 자주 등장하는 3종 세트: ROE 10%+, 부채비율 40~60%, 이익 꾸준.
      예) 오리온: "ROE 10% 이상, 부채비율 20%, PER 7배, PBR 1.1배 => WHY NOT"
          영원무역: "ROE 10%, 부채비율 40%, PBR 0.6배"
    """
    s = (
        0.30 * pct_rank(winsor(df["roe_3y_avg"]))               # 3년 평균 ROE
        + 0.20 * pct_rank(-winsor(df["roe_3y_std"]))            # ROE 변동성 낮을수록 ↑ (꾸준함)
        + 0.20 * pct_rank(winsor(df["op_margin"]))              # 영업이익률
        + 0.20 * pct_rank(-winsor(df["debt_ratio"]))            # 부채비율 낮을수록 ↑
        + 0.10 * pct_rank(winsor(df["rev_cagr_3y"]))            # 매출 우상향
    )
    # 노트: "몇 년간 한 번도 역성장한 적 없다"(서흥) → 무역성장 연속성 가점
    s = s + 0.05 * df["years_no_rev_decline"].clip(0, 5) / 5
    return s


def score_value(df: pd.DataFrame) -> pd.Series:
    """
    가격. 노트의 실제 매수 트리거는 PER 5~9배 / PBR 0.4~1.1배 / 배당 2%+ 구간.
      예) 세아홀딩스 "Forward PER 6배, PBR 0.19배", 메가스터디 "PER 8배 미만, PBR 0.4, 배당 7.55%"
    """
    ep = 1.0 / df["per"].replace([0, np.inf, -np.inf], np.nan)   # 이익수익률
    bp = 1.0 / df["pbr"].replace([0, np.inf, -np.inf], np.nan)
    s = (
        0.40 * pct_rank(winsor(ep))
        + 0.30 * pct_rank(winsor(bp))
        + 0.20 * pct_rank(winsor(df["div_yield"]))
        + 0.10 * pct_rank(winsor(df["fcf_yield"]))
    )
    return s


def score_gap(df: pd.DataFrame) -> pd.Series:
    """
    ★ 핵심 팩터: '실적-주가 괴리'.
    노트 전반의 알파 원천. "돈 꾸준하게 잘 벌고 업황 괜찮은데 주가만 떨어지는 회사.
    CEO가 딱히 할 일 없는 회사들. 가만히 시장만 돌아오면 실적 돌아올 회사들. Overweight!"

    구성: (이익 모멘텀 ≥ 0) AND (주가 모멘텀 << 0) 일수록 고득점.
    단순 낙폭과대와 구분하는 것이 생명 → 이익이 꺾인 종목은 gate로 차단.
    """
    # 이익 게이트: 영업이익 YoY가 -10% 미만이면 괴리가 아니라 '펀더멘털 훼손'
    gate = (df["op_yoy"] > -0.10).astype(float)

    s = (
        0.35 * pct_rank(-winsor(df["ret_12m"]))          # 12개월 수익률 낮을수록 ↑
        + 0.25 * pct_rank(winsor(df["drawdown_52w"]))    # 52주 고점 대비 낙폭 클수록 ↑
        + 0.25 * pct_rank(winsor(df["op_yoy"]))          # 영업이익 개선일수록 ↑
        + 0.15 * pct_rank(winsor(df["rev_yoy"]))
    )
    return s * gate + (1 - gate) * 0.15   # 게이트 탈락 시 사실상 후보 제외


def score_payout(df: pd.DataFrame) -> pd.Series:
    """
    주주환원 여력. 노트: "배당성향 15%도 안되는 걸 보니 배당 확대 가능성 있음"(종근당),
    "배당성향 3% 수준으로 앞으로 많이 올라올 수 있다"(사조대림), 상법개정 테마 반복 언급.
    → '이미 많이 주는 회사'가 아니라 '줄 여력이 있는데 안 주는 회사'가 정답.
    """
    room = (0.50 - df["payout_ratio"]).clip(lower=0)     # 배당성향 50%까지의 여유
    s = (
        0.40 * pct_rank(winsor(room))
        + 0.30 * pct_rank(winsor(df["net_cash_to_mktcap"]))   # 순현금/시총
        + 0.20 * pct_rank(winsor(df["roe_3y_avg"]))           # 여력 있어도 ROE 낮으면 무의미
        + 0.10 * pct_rank(winsor(df["treasury_ratio"]))       # 자사주 보유
    )
    return s


# ═══════════════════════════════════════════════════════════════
# 4. 필터 & 합성
# ═══════════════════════════════════════════════════════════════

def apply_hard_filters(df: pd.DataFrame, cfg: Config = CFG) -> pd.DataFrame:
    m = pd.Series(True, index=df.index)
    reasons = pd.Series("", index=df.index)

    def cut(cond: pd.Series, label: str):
        nonlocal m, reasons
        bad = cond & m
        reasons[bad] = label
        m = m & ~cond

    cut(df["mktcap_eok"] < cfg.min_mktcap_eok, "시총하한")
    cut(df["mktcap_eok"] > cfg.max_mktcap_eok, "초대형제외")
    cut(df["turnover_eok_20d"] < cfg.min_turnover_eok, "유동성")
    cut(df["debt_ratio"] > cfg.max_debt_ratio, "부채과다")
    cut(df["roe_3y_avg"] < cfg.min_roe, "ROE미달")
    if cfg.require_positive_op:
        cut(df["op_ttm"] <= 0, "적자")
    cut(df["ret_3m"] > cfg.max_3m_return, "테마급등")
    cut(df.get("is_admin_issue", pd.Series(False, index=df.index)), "관리종목")

    out = df.copy()
    out["passed"] = m
    out["filter_reason"] = reasons
    return out


def composite(df: pd.DataFrame, cfg: Config = CFG) -> pd.DataFrame:
    d = df.copy()
    d["s_quality"] = score_quality(d)
    d["s_value"] = score_value(d)
    d["s_gap"] = score_gap(d)
    d["s_payout"] = score_payout(d)

    d["score_raw"] = (
        cfg.w_quality * d["s_quality"]
        + cfg.w_value * d["s_value"]
        + cfg.w_gap * d["s_gap"]
        + cfg.w_payout * d["s_payout"]
    )

    d["tilt"] = d["sector"].map(cfg.sector_tilt).fillna(0.0)
    d["score"] = d["score_raw"] + d["tilt"]
    return d.sort_values("score", ascending=False)


# ═══════════════════════════════════════════════════════════════
# 5. 포트폴리오 구성 (비중 상한 + 섹터 상한)
# ═══════════════════════════════════════════════════════════════

def build_portfolio(ranked: pd.DataFrame, cfg: Config = CFG) -> pd.DataFrame:
    cand = ranked[ranked["passed"]].head(cfg.target_positions * 2).copy()

    picks, sector_w, total_w = [], {}, 0.0
    for tkr, row in cand.iterrows():
        if len(picks) >= cfg.target_positions:
            break
        sec = row["sector"]
        # "미분류"(업종 데이터 미연결 상태)는 섹터 상한 적용 제외 —
        # 안 그러면 전종목이 같은 섹터로 묶여 2~3종목에서 멈추는 버그가 생김.
        if sec != "미분류" and sector_w.get(sec, 0.0) >= cfg.max_weight_sector:
            continue
        picks.append(tkr)
        sector_w[sec] = sector_w.get(sec, 0.0) + cfg.max_weight_single

    sel = cand.loc[picks].copy()

    # 스코어 비례 배분 → 단일 10% / 섹터 25% 상한을 동시에 만족할 때까지 반복 재배분
    w = sel["score"].clip(lower=0.01)
    w = w / w.sum()
    sec = sel["sector"]

    for _ in range(200):
        over_single = w > cfg.max_weight_single + 1e-9
        sec_tot = w.groupby(sec).transform("sum")
        # "미분류" 섹터는 상한 재배분 로직에서도 제외 (위와 동일한 이유)
        over_sector = (sec_tot > cfg.max_weight_sector + 1e-9) & (sec != "미분류")
        if not over_single.any() and not over_sector.any():
            break

        if over_single.any():
            excess = (w[over_single] - cfg.max_weight_single).sum()
            w[over_single] = cfg.max_weight_single
            free = ~over_single
        else:
            # 초과 섹터를 상한까지 비례 축소
            scale = (cfg.max_weight_sector / sec_tot).where(over_sector, 1.0)
            excess = (w - w * scale).sum()
            w = w * scale
            free = ~over_sector & (w < cfg.max_weight_single - 1e-9)

        if free.sum() == 0 or excess <= 1e-12:
            break
        w[free] += excess * (w[free] / w[free].sum())

    sel["target_weight"] = w / w.sum()
    return sel


# ═══════════════════════════════════════════════════════════════
# 6. 매매 규칙 엔진 — 보유 종목에 대한 액션 생성
# ═══════════════════════════════════════════════════════════════

@dataclass
class Holding:
    ticker: str
    name: str
    avg_cost: float
    price: float
    weight: float               # 현재 포트 내 비중
    add_downs_done: int = 0
    story_intact: bool = True   # 경쟁우위/실적 스토리 유효 여부 (수동 판단)
    op_yoy: float = 0.0
    ret_3m: float = 0.0
    per_pctile_5y: float = 0.5  # 자기 과거 5년 PER 대비 현재 위치
    cost_recovered: bool = False


def decide(h: Holding, cfg: Config = CFG) -> tuple[str, str]:
    """
    반환: (액션, 근거)
    노트의 매매 원칙을 우선순위대로 적용.
    """
    pnl = h.price / h.avg_cost - 1.0

    # (1) 스토리 훼손 = 유일한 전량매도 사유. 가격 손절은 하지 않는다.
    #     "주가 하락이 매매실패 의미 不" / "펀더멘탈만 잘 보고 밀어붙이고"
    if not h.story_intact:
        return "SELL_ALL", "스토리 훼손 — 가격이 아닌 사실 기반 청산"
    if h.op_yoy < -0.30:
        return "SELL_ALL", f"영업이익 YoY {h.op_yoy:.0%} — 괴리가 아니라 훼손"

    # (2) +100% → 원금 회수, 잔량은 장기보유
    #     "100% 상승 시 원금 회수 후 잔여 물량은 장기 보유"
    if pnl >= cfg.take_profit_recover_cost and not h.cost_recovered:
        return "TRIM_50", "+100% 도달 — 원금 회수 후 잔여분 장기보유"

    # (3) 테마성 급등 → 트림
    #     "인탑스 테마로 급등 → Exit", "삼성SDI ESS 테마 걸려서 확 올랐음. 팔 타임"
    if h.ret_3m > cfg.trim_on_theme_spike:
        return "TRIM_33", f"3개월 +{h.ret_3m:.0%} 급등 — 테마성 과열 구간 일부 실현"

    # (4) 밸류에이션 정상화 → 분할 축소
    #     "국내주식 전반적 저평가 구간 탈출 → 밸류에이션 적정 수준으로 상승"
    if h.per_pctile_5y > 0.80 and pnl > 0.30:
        return "TRIM_33", "자체 5년 밸류 상단 진입 — 분할 축소"

    # (5) 비중 초과 → 리밸런싱
    if h.weight > cfg.max_weight_single * 1.2:
        return "TRIM_TO_CAP", f"비중 {h.weight:.1%} — 단일 10% 상한 초과"

    # (6) 물타기 (스토리 유효 + 하락)
    #     "스토리 살아 있는데 떨어지면 더 사고"
    #     한도 소진 후에도 계속 빠지면 매수가 아니라 재검증. 노트: "물타기 한도" 개념은
    #     "계속 물타면서 나중을 기다리자"의 폭주를 막기 위한 유일한 브레이크.
    limit_level = cfg.add_down_step * cfg.max_add_downs      # 예: -60%
    if h.add_downs_done >= cfg.max_add_downs and pnl <= limit_level:
        return "REVIEW", (f"{pnl:.0%} · 물타기 {h.add_downs_done}회 소진 — "
                          "추가매수 금지, 스토리 재검증")

    next_level = cfg.add_down_step * (h.add_downs_done + 1)  # 1차 -20%, 2차 -40%, 3차 -60%
    if pnl <= next_level and h.add_downs_done < cfg.max_add_downs and h.weight < cfg.max_weight_single:
        return "ADD", f"{pnl:.0%} 하락 · 스토리 유효 · {h.add_downs_done + 1}차 추가매수"

    # (7) 불타기 조건
    #     "오르는데 목표주가가 더 높이 있으면 불 타도 됨"
    if pnl > 0 and h.per_pctile_5y < 0.40 and h.op_yoy > 0.10 and h.weight < cfg.max_weight_single:
        return "ADD", "이익 개선 + 밸류 여전히 하단 — 추격매수 허용"

    return "HOLD", "조건 미충족 — 보유 유지"


# ═══════════════════════════════════════════════════════════════
# 7. 실데이터 어댑터 (pykrx + OpenDartReader)
# ═══════════════════════════════════════════════════════════════

def load_real(date: str, bsns_year: int = 2025) -> pd.DataFrame:
    """
    KRX 공식 Open API로 가격·시가총액 지표(KOSPI), data_pipeline.py 캐시로 재무 지표를 조립.
    사전 준비:
      1) setx KRX_API_KEY "..."  / setx DART_API_KEY "..." 등록
      2) python data_pipeline.py --build --year 2025 --date 20260807  실행해서
         .cache/finance.parquet 만들어둘 것
    """
    from data_pipeline import FINANCE_CACHE, get_kospi_universe

    if not FINANCE_CACHE.exists():
        raise RuntimeError(
            "재무 캐시가 없습니다. 먼저 실행하세요: "
            "python data_pipeline.py --build --year 2025 --date 20260807"
        )

    # ---- 가격·시가총액 (KRX 공식 API, 유가증권 일별매매정보) ----
    df = get_kospi_universe(date)   # index=stock_code, columns: name, close, fluc_rt, trdval, mktcap, ...
    df["mktcap_eok"] = df["mktcap"] / 1e8
    df["turnover_eok_20d"] = df["trdval"] / 1e8   # TODO: 20일 평균으로 교체 (현재는 당일 값)
    df["sector"] = "미분류"  # TODO: 종목기본정보 API(SECT_TP_NM)로 업종 채워 넣을 것
    df["ret_3m"] = np.nan   # TODO: 과거 시점 기준일로 한 번 더 조회해서 등락률 계산
    df["ret_12m"] = np.nan  # TODO: 위와 동일
    df["drawdown_52w"] = np.nan  # TODO: 일별 시세 히스토리 필요 (아래 보완 계획 참고)

    # ---- 밸류/재무 캐시 merge ----
    df["div_yield"] = np.nan    # TODO: KRX API에 배당수익률 서비스가 없어 비워둠. 필요시 DART 배당 공시로 보완
    df["payout_ratio"] = np.nan # TODO: 위와 동일
    df["fcf_yield"] = np.nan
    df["net_cash_to_mktcap"] = np.nan
    df["treasury_ratio"] = np.nan
    df["is_admin_issue"] = False

    fin = pd.read_parquet(FINANCE_CACHE)
    fin = fin[fin["bsns_year"] == bsns_year].set_index("stock_code")
    df = df.join(fin, how="inner")

    # PER/PBR은 "시가총액 ÷ 재무값" 방식으로 계산 (1주당 EPS/BPS를 따로 구할 필요 없음)
    #   PER = 시가총액 / 당기순이익  (= 주가 / EPS 와 수학적으로 동일)
    #   PBR = 시가총액 / 자본총계    (= 주가 / BPS 와 수학적으로 동일)
    df["per"] = df["mktcap"] / df["net_income_ttm"].where(df["net_income_ttm"] > 0)
    df["pbr"] = df["mktcap"] / df["total_equity"].where(df["total_equity"] > 0)

    return df


# ═══════════════════════════════════════════════════════════════
# 8. 데모 — 합성 데이터로 로직 검증
# ═══════════════════════════════════════════════════════════════

def make_demo(n: int = 300, seed: int = 7) -> pd.DataFrame:
    rng = np.random.default_rng(seed)
    sectors = list(CFG.sector_tilt.keys()) + ["기계", "화학", "IT서비스", "유통"]
    df = pd.DataFrame(index=[f"{i:06d}" for i in range(n)])
    df["name"] = [f"종목{i}" for i in range(n)]
    df["sector"] = rng.choice(sectors, n)
    df["mktcap_eok"] = rng.lognormal(7.5, 1.3, n)
    df["turnover_eok_20d"] = rng.lognormal(1.5, 1.2, n)
    df["roe_3y_avg"] = rng.normal(8, 6, n)
    df["roe_3y_std"] = np.abs(rng.normal(4, 2.5, n))
    df["op_margin"] = rng.normal(7, 6, n)
    df["debt_ratio"] = np.abs(rng.normal(90, 60, n))
    df["rev_cagr_3y"] = rng.normal(0.04, 0.10, n)
    df["years_no_rev_decline"] = rng.integers(0, 6, n)
    df["per"] = np.abs(rng.lognormal(2.3, 0.7, n))
    df["pbr"] = np.abs(rng.lognormal(-0.2, 0.6, n))
    df["div_yield"] = np.abs(rng.normal(0.018, 0.015, n))
    df["fcf_yield"] = rng.normal(0.04, 0.05, n)
    df["ret_12m"] = rng.normal(0.05, 0.35, n)
    df["ret_3m"] = rng.normal(0.02, 0.20, n)
    df["drawdown_52w"] = np.abs(rng.normal(0.25, 0.15, n))
    df["op_yoy"] = rng.normal(0.05, 0.30, n)
    df["rev_yoy"] = rng.normal(0.04, 0.15, n)
    df["op_ttm"] = rng.normal(200, 400, n)
    df["payout_ratio"] = np.abs(rng.normal(0.20, 0.15, n))
    df["net_cash_to_mktcap"] = rng.normal(0.05, 0.20, n)
    df["treasury_ratio"] = np.abs(rng.normal(0.02, 0.03, n))
    df["is_admin_issue"] = rng.random(n) < 0.02

    # '서흥형' 종목 심기: 체력 좋은데 주가만 빠진 케이스
    for i in [3, 17, 42, 88, 150]:
        t = df.index[i]
        df.loc[t, ["roe_3y_avg", "roe_3y_std", "op_margin", "debt_ratio"]] = [13, 2.0, 11, 42]
        df.loc[t, ["per", "pbr", "div_yield"]] = [6.5, 0.55, 0.028]
        df.loc[t, ["ret_12m", "ret_3m", "drawdown_52w"]] = [-0.38, -0.09, 0.46]
        df.loc[t, ["op_yoy", "rev_yoy", "payout_ratio"]] = [0.06, 0.05, 0.09]
        df.loc[t, ["mktcap_eok", "turnover_eok_20d", "op_ttm"]] = [4200, 25, 480]
        df.loc[t, "is_admin_issue"] = False
        df.loc[t, "sector"] = ["건강기능식품", "음식료", "화장품", "반도체소재", "조선기자재"][
            [3, 17, 42, 88, 150].index(i)
        ]
        df.loc[t, "name"] = f"★괴리형{i}"
    return df


def run_demo():
    pd.set_option("display.width", 200, "display.max_columns", 50)
    df = make_demo()
    filt = apply_hard_filters(df)
    ranked = composite(filt)

    print("=" * 78)
    print("STEP 1 — 하드 필터")
    print("=" * 78)
    print(f"유니버스 {len(df)} → 통과 {int(filt['passed'].sum())}")
    print(filt.loc[~filt["passed"], "filter_reason"].value_counts().to_string())

    print("\n" + "=" * 78)
    print("STEP 2 — 종합 랭킹 상위 15")
    print("=" * 78)
    cols = ["name", "sector", "per", "pbr", "roe_3y_avg", "debt_ratio",
            "ret_12m", "op_yoy", "s_quality", "s_value", "s_gap", "s_payout", "score"]
    top = ranked[ranked["passed"]].head(15)[cols]
    print(top.round(3).to_string())

    print("\n" + "=" * 78)
    print("STEP 3 — 포트폴리오 (단일 10% / 섹터 25% 상한)")
    print("=" * 78)
    pf = build_portfolio(ranked)
    print(pf[["name", "sector", "score", "target_weight"]].round(4).to_string())
    print(f"\n합계 비중: {pf['target_weight'].sum():.1%} | 종목수: {len(pf)}")
    print("섹터별:\n" + pf.groupby("sector")["target_weight"].sum().round(3).to_string())

    print("\n" + "=" * 78)
    print("STEP 4 — 보유종목 매매 규칙 엔진")
    print("=" * 78)
    cases = [
        Holding("001", "서흥형(장기물림)", 30000, 21000, 0.05, add_downs_done=1, op_yoy=0.05),
        Holding("002", "서흥형(폭등후)", 21000, 43000, 0.09, op_yoy=0.20, ret_3m=0.72),
        Holding("003", "테마급등형", 10000, 17000, 0.06, op_yoy=0.02, ret_3m=0.68),
        Holding("004", "펀더멘털훼손", 50000, 38000, 0.04, op_yoy=-0.42),
        Holding("005", "비중초과", 8000, 14000, 0.14, op_yoy=0.08, per_pctile_5y=0.5),
        Holding("006", "물타기한도소진", 60000, 18000, 0.09, add_downs_done=3, op_yoy=-0.05),
        Holding("007", "이익개선+저밸류", 12000, 14500, 0.05, op_yoy=0.18, per_pctile_5y=0.28),
        Holding("008", "밸류상단도달", 20000, 29000, 0.06, op_yoy=0.06, per_pctile_5y=0.88),
    ]
    for h in cases:
        act, why = decide(h)
        pnl = h.price / h.avg_cost - 1
        print(f"{h.name:<16} 손익{pnl:+7.1%}  →  {act:<14} | {why}")


def run_real(date: str, bsns_year: int, top_n: int, export: str | None,
             export_json: str | None = None):
    """
    종목 스크리닝 전용 실행기. (포트폴리오 비중 구성은 하지 않음 —
    이건 '어떤 종목을 볼지' 후보를 추리는 용도이지, 실제 매매 비중 결정용이 아님)

    기준 시점:
      - 가격/시가총액: date 인자 기준 (예: 20260807 = 2026년 8월 7일 종가)
      - 재무데이터(ROE·부채비율 등): bsns_year 기준 사업보고서 (예: 2025 = 2025년 사업연도 재무제표)
      두 시점이 다를 수 있으니 참고할 것.
    """
    pd.set_option("display.width", 200, "display.max_columns", 50)
    d = load_real(date, bsns_year=bsns_year)
    filt = apply_hard_filters(d)
    ranked = composite(filt)

    print("=" * 78)
    print(f"STEP 1 — 하드 필터 (가격기준일 {date} / 재무기준연도 {bsns_year})")
    print("=" * 78)
    print(f"유니버스 {len(d)} → 통과 {int(filt['passed'].sum())}")
    if (~filt["passed"]).any():
        print(filt.loc[~filt["passed"], "filter_reason"].value_counts().to_string())

    print("\n" + "=" * 78)
    print(f"STEP 2 — 종합 랭킹 상위 {top_n}  (점수 높은 순 = 4대 팩터 종합 매력도)")
    print("=" * 78)
    cols = ["name", "sector_raw", "close", "per", "pbr", "roe_3y_avg", "debt_ratio",
            "ret_12m", "op_yoy", "s_quality", "s_value", "s_gap", "s_payout", "score"]
    cols = [c for c in cols if c in ranked.columns]
    if "mktcap_eok" in ranked.columns:
        cols = cols[:2] + ["mktcap_eok"] + cols[2:]  # name, sector_raw 다음에 시가총액 삽입

    top = ranked[ranked["passed"]].head(top_n)[cols]
    print(top.round(3).to_string())

    KOR_NAMES = {
        "name": "종목명", "sector_raw": "시장", "mktcap_eok": "시가총액(억)",
        "close": "종가", "per": "PER", "pbr": "PBR", "roe_3y_avg": "ROE(3년평균%)",
        "debt_ratio": "부채비율(%)", "ret_12m": "12개월수익률", "op_yoy": "영업이익YoY",
        "s_quality": "체력점수", "s_value": "가격점수", "s_gap": "괴리점수",
        "s_payout": "환원여력점수", "score": "종합점수",
    }

    if export:
        screened = top.rename(columns=KOR_NAMES).round(3)
        screened.index.name = "종목코드"

        def _write(path):
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                screened.to_excel(writer, sheet_name="스크리닝결과", index=True)
                filt[["passed", "filter_reason"]].to_excel(writer, sheet_name="필터현황", index=True)

                from openpyxl.utils import get_column_letter
                ws = writer.sheets["스크리닝결과"]
                ws.freeze_panes = "A2"  # 헤더 행 고정
                for i, col in enumerate([screened.index.name] + list(screened.columns), start=1):
                    width = max(10, min(20, len(str(col)) * 2 + 4))
                    ws.column_dimensions[get_column_letter(i)].width = width
                for cell in ws[1]:
                    cell.font = cell.font.copy(bold=True)

        try:
            _write(export)
            print(f"\n[export] 엑셀 저장 완료 → {export} (스크리닝결과 시트에 상위 {len(screened)}종목만 정리됨)")
        except PermissionError:
            fallback = export.replace(".xlsx", f"_{pd.Timestamp.now().strftime('%H%M%S')}.xlsx")
            print(f"\n[export 실패] '{export}' 파일이 이미 열려 있는 것 같습니다. "
                  f"엑셀에서 닫고 다시 시도하거나, 새 이름으로 저장합니다.")
            _write(fallback)
            print(f"[export] 대신 저장 완료 → {fallback}")
        except ImportError:
            print("\n[export 실패] openpyxl이 설치되어 있지 않습니다. "
                  "터미널에서 'py -m pip install openpyxl' 실행 후 다시 시도하세요.")

    if export_json:
        import json
        from pathlib import Path as _Path

        records = []
        for code, row in top.iterrows():
            rec = {"stock_code": str(code)}
            for c in cols:
                v = row[c]
                if pd.isna(v):
                    rec[c] = None
                elif isinstance(v, (int, float, np.floating, np.integer)):
                    rec[c] = round(float(v), 4)
                else:
                    rec[c] = str(v)
            records.append(rec)

        payload = {
            "as_of_date": date,
            "financial_year": bsns_year,
            "generated_at": pd.Timestamp.now("UTC").isoformat(),
            "universe_total": int(len(d)),
            "universe_passed": int(filt["passed"].sum()),
            "columns": cols,
            "column_labels_ko": {c: KOR_NAMES.get(c, c) for c in cols},
            "results": records,
        }

        out_path = _Path(export_json)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"\n[export] JSON 저장 완료 → {export_json} ({len(records)}종목)")

    return ranked


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--demo", action="store_true")
    ap.add_argument("--run", action="store_true")
    ap.add_argument("--date", default="20260810")
    ap.add_argument("--year", type=int, default=2025, help="재무캐시 사업연도")
    ap.add_argument("--top", type=int, default=60, help="화면·엑셀에 보여줄 상위 종목 수")
    ap.add_argument("--export", default="screening_result.xlsx", help="엑셀로 저장할 파일명 (빈 문자열이면 저장 안 함)")
    ap.add_argument("--export-json", default="", help="웹사이트용 JSON 저장 경로 (예: web/data/results.json)")
    a = ap.parse_args()
    if a.run:
        run_real(a.date, a.year, a.top, a.export if a.export else None,
                  a.export_json if a.export_json else None)
    else:
        run_demo()
